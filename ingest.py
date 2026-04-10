import os
import fitz  # PyMuPDF
import openpyxl
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import Chroma
from langchain_huggingface import HuggingFaceEmbeddings

DOCS_DIR = "./documents"       # mets tes PDFs et Excel ici
CHROMA_DIR = "./chroma_db"     # la base vectorielle sera créée ici

def extract_pdf(path: str) -> str:
    doc = fitz.open(path)
    text = ""
    for i, page in enumerate(doc):
        text += f"\n[Page {i+1}]\n" + page.get_text()
    return text

def extract_excel(path: str) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text += f"\n[Feuille: {sheet}]\n"
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row)
            if row_text.strip():
                text += row_text + "\n"
    return text

def load_all_documents(directory: str) -> list[dict]:
    docs = []
    for root, dirs, files in os.walk(directory):  # os.walk descend dans les sous-dossiers
        for filename in files:
            filepath = os.path.join(root, filename)
            if filename.endswith(".pdf"):
                print(f"  Lecture PDF : {filepath}")
                content = extract_pdf(filepath)
                docs.append({"content": content, "source": filename})
            elif filename.endswith((".xlsx", ".xls")):
                print(f"  Lecture Excel : {filepath}")
                content = extract_excel(filepath)
                docs.append({"content": content, "source": filename})
    return docs

def ingest():
    print("Chargement des documents...")
    raw_docs = load_all_documents(DOCS_DIR)

    # Chunking : 500 tokens, 50 de chevauchement
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=500,
        chunk_overlap=50,
        separators=["\n\n", "\n", ".", " "]
    )

    all_chunks = []
    all_metadatas = []

    for doc in raw_docs:
        chunks = splitter.split_text(doc["content"])
        for chunk in chunks:
            all_chunks.append(chunk)
            all_metadatas.append({"source": doc["source"]})

    print(f"{len(all_chunks)} chunks créés. Indexation en cours...")

    embeddings = HuggingFaceEmbeddings(model_name="all-MiniLM-L6-v2")

    vectorstore = Chroma.from_texts(
        texts=all_chunks,
        embedding=embeddings,
        metadatas=all_metadatas,
        persist_directory=CHROMA_DIR
    )

    print("Indexation terminée. Base vectorielle sauvegardée dans ./chroma_db")

if __name__ == "__main__":
    ingest()
